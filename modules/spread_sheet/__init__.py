from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from typing import List
import os


class ExcelHandler:
    def __init__(self, filename: str, sheetname: str):
        self.filename = filename
        self.sheetname = sheetname

        if os.path.exists(filename):
            self.workbook = load_workbook(filename)
            
            if sheetname in self.workbook.sheetnames:
                self.sheet = self.workbook[sheetname]
            else:
                self.sheet = self.workbook.create_sheet(title=sheetname)
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = sheetname
        self.bold_font = Font(bold=True)

    def insert_header(self, headers: List[str]):
        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.bold_font

    def insert_data(self, data: List[List[str]]):
        for row, product in enumerate(data, start=2):
            for col, value in enumerate(product, start=1):
                cell = self.sheet.cell(row=row, column=col)
                cell.value = value

    def save(self):
        self.workbook.save(filename=self.filename)
